import pandas as pd
from sqlalchemy import create_engine
from openpyxl.utils import get_column_letter
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def regenerate_excel_from_db():
    logger = logging.getLogger("regenerate_excel")
    db_path = 'data/cases.db'
    excel_path = 'data/cases.xlsx'
    
    logger.info("开始重新生成 Excel 文件...")
    
    engine = create_engine(f'sqlite:///{db_path}')
    
    # 修改 SQL 查询，增加三个新列
    df = pd.read_sql_query(
        """
        SELECT 
            case_name, 
            notice_start_date, 
            notice_end_date, 
            source_url, 
            region, 
            created_at,
            参与集中的经营者,
            审结时间,
            是否已匹配
        FROM cases 
        ORDER BY notice_start_date DESC
        """, 
        engine
    )
    
    # 处理日期列
    for col in ['notice_start_date', 'notice_end_date', 'created_at', '审结时间']:
        df[col] = df[col].str.slice(0, 10) if df[col].dtype == 'object' else df[col]
        df[col] = pd.to_datetime(df[col], format='%Y-%m-%d', errors='coerce')
    
    # 定义列名映射（增加新列）
    column_mapping = {
        'case_name': '案件名称',
        'notice_start_date': '公示开始日期',
        'notice_end_date': '公示结束日期',
        'source_url': '来源链接',
        'region': '地区',
        'created_at': '爬取时间',
        '参与集中的经营者': '参与集中的经营者',
        '审结时间': '审结时间',
        '是否已匹配': '是否已匹配'
    }
    
    # 重命名列
    df = df.rename(columns=column_mapping)
    
    # 写入Excel
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='案件列表')
        worksheet = writer.sheets['案件列表']
        
        # 设置列宽（增加新列的宽度设置）
        for idx, column in enumerate(df.columns, 1):
            if column == '案件名称':
                width = 60
            elif column in ['公示开始日期', '公示结束日期', '爬取时间', '审结时间']:
                width = 15
            elif column == '来源链接':
                width = 50
            elif column == '地区':
                width = 10
            elif column == '参与集中的经营者':
                width = 50  # 参与者名称可能较长
            elif column == '是否已匹配':
                width = 10
            else:
                width = 15
                
            worksheet.column_dimensions[get_column_letter(idx)].width = width
            
        # 设置日期格式
        date_format = 'yyyy-mm-dd'
        for col in ['公示开始日期', '公示结束日期', '爬取时间', '审结时间']:
            if col in df.columns:
                col_idx = df.columns.get_loc(col) + 1
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=col_idx)
                    cell.number_format = date_format
    
    logger.info(f"Excel 文件生成完成，共 {len(df)} 条记录")

if __name__ == "__main__":
    regenerate_excel_from_db()
